# routers/import_export.py

from fastapi import APIRouter, Form, Query, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse, Response
from typing import List, Dict, Optional
from datetime import datetime
import io
import csv
import openpyxl

from backup_utils import (
    PORTFOLIOS,
    PORTFOLIO_HISTORY,
    DEFAULT_PORTFOLIO,
    rebuild_portfolio_history_from_transactions,
)
from core.state import resolve_portfolio  # your shared resolve helper
from core.state import TAGS  # tags dict shared with UI

router = APIRouter()


@router.post("/importcsv")
async def importcsv(
    portfolioname: str = Form(...),
    file: UploadFile = File(...),
):
    pname = portfolioname.strip() or DEFAULT_PORTFOLIO

    content = await file.read()
    if not content:
        return RedirectResponse(
            url=f"/?portfolio={pname}&error=Uploaded CSV is empty.",
            status_code=303,
        )

    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))

    required_cols = [
        "Symbol",
        "Current Price",
        "Trade Date",
        "Purchase Price",
        "Quantity",
    ]
    if not reader.fieldnames:
        return RedirectResponse(
            url=f"/?portfolio={pname}&error=CSV has no header row.",
            status_code=303,
        )
    for col in required_cols:
        if col not in reader.fieldnames:
            return RedirectResponse(
                url=f"/?portfolio={pname}&error=CSV missing required column {col}",
                status_code=303,
            )

    if pname not in PORTFOLIOS:
        PORTFOLIOS[pname] = {
            "tickers": [],
            "positions": {},
            "transactions": [],
        }

    pf = PORTFOLIOS[pname]
    positions: Dict[str, Dict[str, float]] = {}
    transactions: List[Dict] = []
    PORTFOLIO_HISTORY[pname] = []
    TAGS[pname] = {}

    for row in reader:
        symbol = (row["Symbol"] or "").strip().upper()
        if not symbol:
            continue

        tradedate_raw = (row["Trade Date"] or "").strip()
        buyprice = float(row["Purchase Price"] or 0.0)
        qty = float(row["Quantity"] or 0.0)

        if not tradedate_raw or len(tradedate_raw) != 8:
            tradedt = datetime.now()
        else:
            try:
                tradedt = datetime.strptime(tradedate_raw, "%Y%m%d")
            except Exception:
                tradedt = datetime.now()

        pos = positions.get(symbol, {"qty": 0.0, "buy": 0.0})
        oldqty = pos["qty"]
        oldbuy = pos["buy"]

        newqty = oldqty + qty
        if newqty != 0:
            if oldqty != 0:
                newbuy = (oldbuy * oldqty + buyprice * qty) / newqty
            else:
                newbuy = buyprice
        else:
            newbuy = buyprice

        pos["qty"] = newqty
        pos["buy"] = newbuy
        positions[symbol] = pos

        transactions.append(
            {
                "time": tradedt.strftime("%Y-%m-%d %H:%M:%S"),
                "portfolio": pname,
                "symbol": symbol,
                "action": "BUY",
                "qty": qty,
                "price": buyprice,
                "note": f"Imported from CSV at {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            }
        )

    pf["tickers"] = sorted(positions.keys())
    pf["positions"] = positions
    pf["transactions"] = transactions
    rebuild_portfolio_history_from_transactions(pname)

    return RedirectResponse(url=f"/?portfolio={pname}", status_code=303)


@router.post("/importcsvmulti")
async def importcsvmulti(
    portfolionames: str = Form(...),
    files: List[UploadFile] = File(...),
):
    tokens = [t.strip() for t in portfolionames.split(",")]
    names = [t for t in tokens if t]

    if not names:
        return RedirectResponse(
            url="/?error=No portfolio names provided for bulk CSV import.",
            status_code=303,
        )
    if len(names) != len(files):
        return RedirectResponse(
            url="/?error=Number of portfolio names and CSV files must match.",
            status_code=303,
        )

    for pname, file in zip(names, files):
        pname = pname or DEFAULT_PORTFOLIO
        content = await file.read()
        if not content:
            continue

        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))

        required_cols = [
            "Symbol",
            "Current Price",
            "Trade Date",
            "Purchase Price",
            "Quantity",
        ]
        if not reader.fieldnames or any(col not in reader.fieldnames for col in required_cols):
            # Skip malformed CSV, do not stop whole batch
            continue

        if pname not in PORTFOLIOS:
            PORTFOLIOS[pname] = {
                "tickers": [],
                "positions": {},
                "transactions": [],
            }

        pf = PORTFOLIOS[pname]
        positions: Dict[str, Dict[str, float]] = {}
        transactions: List[Dict] = []
        PORTFOLIO_HISTORY[pname] = []
        TAGS[pname] = {}

        for row in reader:
            symbol = (row["Symbol"] or "").strip().upper()
            if not symbol:
                continue

            tradedate_raw = (row["Trade Date"] or "").strip()
            buyprice = float(row["Purchase Price"] or 0.0)
            qty = float(row["Quantity"] or 0.0)

            if not tradedate_raw or len(tradedate_raw) != 8:
                tradedt = datetime.now()
            else:
                try:
                    tradedt = datetime.strptime(tradedate_raw, "%Y%m%d")
                except Exception:
                    tradedt = datetime.now()

            pos = positions.get(symbol, {"qty": 0.0, "buy": 0.0})
            oldqty = pos["qty"]
            oldbuy = pos["buy"]

            newqty = oldqty + qty
            if newqty != 0:
                if oldqty != 0:
                    newbuy = (oldbuy * oldqty + buyprice * qty) / newqty
                else:
                    newbuy = buyprice
            else:
                newbuy = buyprice

            pos["qty"] = newqty
            pos["buy"] = newbuy
            positions[symbol] = pos

            transactions.append(
                {
                    "time": tradedt.strftime("%Y-%m-%d %H:%M:%S"),
                    "portfolio": pname,
                    "symbol": symbol,
                    "action": "BUY",
                    "qty": qty,
                    "price": buyprice,
                    "note": f"Imported from CSV bulk at {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                }
            )

        pf["tickers"] = sorted(positions.keys())
        pf["positions"] = positions
        pf["transactions"] = transactions
        rebuild_portfolio_history_from_transactions(pname)

    # After bulk import, redirect to the last portfolio in the list
    return RedirectResponse(url=f"/?portfolio={names[-1]}", status_code=303)


@router.get("/portfoliodownload")
async def portfoliodownload(portfolio: str = Query(...)):
    pname = resolve_portfolio(portfolio)
    pf = PORTFOLIOS[pname]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio"
    ws.append(["Symbol", "Current Price", "Trade Date", "Purchase Price", "Quantity"])

    today = datetime.now().strftime("%Y%m%d")
    for sym in pf["tickers"]:
        pos = pf["positions"].get(sym)
        if not pos:
            continue
        ws.append(
            [
                sym,
                "",
                today,
                pos.get("buy"),
                pos.get("qty"),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{pname.replace(' ', '_')}_portfolio.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/portfolioupload")
async def portfolioupload(
    portfolio: str = Form(...),
    file: UploadFile = File(...),
):
    pname = portfolio.strip() or DEFAULT_PORTFOLIO

    content = await file.read()
    if not content:
        return RedirectResponse(
            url=f"/?portfolio={pname}&error=Uploaded XLSX is empty.",
            status_code=303,
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        return RedirectResponse(
            url=f"/?portfolio={pname}&error=Failed to read XLSX file.",
            status_code=303,
        )

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    colmap = {name: i for i, name in enumerate(header)}

    for col in ["Symbol", "Purchase Price", "Quantity"]:
        if col not in colmap:
            return RedirectResponse(
                url=f"/?portfolio={pname}&error=XLSX missing required column {col}",
                status_code=303,
            )

    if pname not in PORTFOLIOS:
        PORTFOLIOS[pname] = {
            "tickers": [],
            "positions": {},
            "transactions": [],
        }

    pf = PORTFOLIOS[pname]
    positions: Dict[str, Dict[str, float]] = {}
    tickers: List[str] = []
    transactions: List[Dict] = []
    TAGS[pname] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol = str(row[colmap["Symbol"]] or "").strip().upper()
        if not symbol:
            continue
        buy = float(row[colmap["Purchase Price"]] or 0.0)
        qty = float(row[colmap["Quantity"]] or 0.0)

        pos = positions.get(symbol, {"qty": 0.0, "buy": 0.0})
        oldqty = pos["qty"]
        oldbuy = pos["buy"]

        newqty = oldqty + qty
        if newqty != 0:
            if oldqty != 0:
                newbuy = (oldbuy * oldqty + buy * qty) / newqty
            else:
                newbuy = buy
        else:
            newbuy = buy

        pos["qty"] = newqty
        pos["buy"] = newbuy
        positions[symbol] = pos
        tickers.append(symbol)

        transactions.append(
            {
                "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "portfolio": pname,
                "symbol": symbol,
                "action": "BUY",
                "qty": qty,
                "price": buy,
                "note": "Imported from XLSX",
            }
        )

    pf["tickers"] = sorted(set(tickers))
    pf["positions"] = positions
    pf["transactions"].extend(transactions)
    rebuild_portfolio_history_from_transactions(pname)

    return RedirectResponse(url=f"/?portfolio={pname}", status_code=303)


@router.get("/portfolioexportcsv")
async def portfolioexportcsv(portfolio: str = Query(...)):
    pname = resolve_portfolio(portfolio)
    pf = PORTFOLIOS[pname]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ["Symbol", "Current Price", "Trade Date", "Purchase Price", "Quantity"]
    )
    today = datetime.now().strftime("%Y%m%d")

    for sym in pf["tickers"]:
        pos = pf["positions"].get(sym)
        if not pos:
            continue
        writer.writerow(
            [
                sym,
                "",
                today,
                pos.get("buy"),
                pos.get("qty"),
            ]
        )

    buf.seek(0)
    filename = f"{pname.replace(' ', '_')}_portfolio.csv"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(content=buf.getvalue(), media_type="text/csv", headers=headers)

